from flask import Flask, render_template, request
import random
import re
import os
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.ensemble import RandomForestClassifier
from sklearn.metrics import accuracy_score, precision_score, recall_score, f1_score
import pickle
import openpyxl
from nltk.corpus import stopwords
from nltk.stem import PorterStemmer, WordNetLemmatizer
from nltk.tokenize import word_tokenize
import enchant
import nltk
from spellchecker import SpellChecker
from flask_pymongo import PyMongo
from bson.objectid import ObjectId

app = Flask(__name__)
app.config['MONGO_URI'] = 'mongodb://localhost:27017/mydataset'  # Replace with your MongoDB connection URI
mongo = PyMongo(app)

# Load the training data from an Excel file
data = pd.read_excel("C:/Users/varsh/Downloads/Locus-for-Local-Retailers-main/Locus-for-Local-Retailers-main\Request Response 50.xlsx")
excel_file = pd.read_excel("C:/Users/varsh/Downloads/Locus-for-Local-Retailers-main/Locus-for-Local-Retailers-main/DATA_FINAL_SHOPS.xlsx")
sheet_name = 'Sheet1'  # Replace with the desired sheet name

# Split the data into training and testing sets
X_train, X_test, y_train, y_test = train_test_split(data['Keywords'], data['Response'], test_size=0.2, random_state=42)

# Convert the request data into numerical features using TF-IDF vectorization
vectorizer = TfidfVectorizer()
X_train = vectorizer.fit_transform(X_train)
X_test = vectorizer.transform(X_test)

# Train a machine learning model on the vectorized data
model = RandomForestClassifier()
model.fit(X_train, y_train)
y_pred = model.predict(X_test)

# Calculate the metrics
accuracy = accuracy_score(y_test, y_pred)
precision = precision_score(y_test, y_pred, average='weighted')
recall = recall_score(y_test, y_pred, average='weighted')
f1 = f1_score(y_test, y_pred, average='weighted')

# Save the vectorizer and model to disk for later use
with open('vectorizer.pkl', 'wb') as file:
    pickle.dump(vectorizer, file)
with open('model.pkl', 'wb') as file:
    pickle.dump(model, file)

# Load the vectorizer and model from disk
with open('vectorizer.pkl', 'rb') as file:
    vectorizer = pickle.load(file)
with open('model.pkl', 'rb') as file:
    model = pickle.load(file)


def link(rese, url):
    text = rese
    hyperlink = '<a href="{}">{}</a>'.format(url, text)
    return hyperlink


def generate_response(s):
    if s == "":
        return "Please enter a valid input"
    try:
        # Convert the input text to a vector using the vectorizer
        vectorized_text = vectorizer.transform([s])
        # Use the trained model to make a prediction
        prediction = model.predict(vectorized_text)
        # Get the corresponding response from the Excel file
        response = data.loc[data['Keywords'] == s, 'Response'].values
        if len(response) == 0:
            return prediction[0], "Sorry, still under construction."
        else:
            return prediction[0], response[0]
    except Exception as e:
        print(f"Error during model prediction: {e}")
        return "", "Sorry, we encountered an error while processing your request."


def display_matching_rows(excel_file, sheet_name, target_row_name):
    try:
        print("s: ", target_row_name)
        workbook = openpyxl.load_workbook(excel_file)
        worksheet = workbook[sheet_name]
        matching_rows = []
        ans = '\n Please login into your Instagram and then access the link:) \n'
        cnt = 1
        row_cnt = 0
        for row in worksheet.iter_rows(values_only=True):
            row_cnt += 1
            if row[0] == target_row_name:
                row_number = row_cnt
                data = worksheet.cell(row=row_number, column=2).hyperlink.target
                ans = ans + str(cnt) + ".) " + row[1] + ' - ' + data + ', ' + os.linesep
                cnt += 1
        return ans
    except Exception as e:
        print(f"Error in reading Excel file: {e}")
        return "Sorry, could not find relevant information."


# Text cleaning and spell checking functions remain the same...

@app.route('/')
def index():
    return render_template('home.html')


@app.route('/loginAc', methods=['POST'])
def loginAc():
    try:
        if request.method == 'POST':
            username = request.form['username']
            password = request.form['password']
            collection = mongo.db.User
            cursor = collection.find_one({"username": username, "password": password})
            if cursor:
                return render_template('newcbotline.html')
            else:
                return render_template('home.html', err='Invalid credentials')
    except Exception as e:
        print(f"Error during login: {e}")
        return render_template('home.html', err='An error occurred during login.')


@app.route('/chat', methods=['POST'])
def chat():
    try:
        global user_input_list
        input_text = request.form['user_input']
        end_greetings = ['bye', 'goodbye', 'see you later', 'see ya', 'adios']
        greeting_keywords = ["hello", "hi", "greetings", "hey", "nice to meet you", 'hii']
        greetings = ["Hello! How can I assist you today?", ...]

        if input_text.lower() in greeting_keywords:
            response = random.choice(greetings)
        elif input_text.lower() in end_greetings:
            response = "Thank you for using our chatbot. Please provide feedback. Visit again!"
        else:
            user_input_list.append(input_text)
            input_txt = ' '.join(user_input_list)
            input_txtt = nlp_preprocessing(input_txt)
            s = ",".join(input_txtt)
            if not s:
                return "Please enter a valid input"
            label, response = generate_response(s)

        if 'shops' in response:
            list_for_shops = [...]  # predefined list
            target_row_name = ""
            for word in list_for_shops:
                if word in input_txtt:
                    target_row_name += word + ','
            response1 = display_matching_rows("C:/path_to/DATA_FINAL_SHOPS.xlsx", sheet_name, target_row_name.rstrip(','))
            response = response + os.linesep + response1

        # Store chat history in MongoDB
        mongo.db.chat_history.insert_one({'user_input': input_text, 'response': response})

        return response

    except Exception as e:
        print(f"Error during chat processing: {e}")
        return "Sorry, an error occurred while processing your request."


if __name__ == '__main__':
    app.run(debug=False, port=5002)
